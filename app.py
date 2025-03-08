import os
import certifi
import secrets

# Función auxiliar para detectar números flotantes\ndef is_float(value):\n    try:\n        float(value)\n        return True\n    except (ValueError, TypeError):\n        return False
from datetime import datetime, timedelta
import tempfile
import zipfile
from dotenv import load_dotenv
import sys

# Cargar variables de entorno desde .env
load_dotenv()

# Validación de variables de entorno críticas
variables_criticas = {
    'FLASK_SECRET_KEY': 'clave secreta para Flask',
    'MONGO_URI': 'URI de conexión a MongoDB',
    'MAIL_SERVER': 'servidor de correo SMTP',
    'MAIL_PORT': 'puerto del servidor de correo',
    'MAIL_USERNAME': 'nombre de usuario del correo',
    'MAIL_PASSWORD': 'contraseña del correo',
    'MAIL_DEFAULT_SENDER_NAME': 'nombre del remitente por defecto',
    'MAIL_DEFAULT_SENDER_EMAIL': 'correo del remitente por defecto'
}

variables_faltantes = []
for variable, descripcion in variables_criticas.items():
    if not os.environ.get(variable):
        variables_faltantes.append(f"{variable} ({descripcion})")

if variables_faltantes:
    error_message = "ERROR: Faltan las siguientes variables de entorno críticas:\n"
    for var in variables_faltantes:
        error_message += f"  - {var}\n"
    error_message += "\nPor favor, configura estas variables en el archivo .env antes de ejecutar la aplicación."
    
    print(error_message, file=sys.stderr)
    # Opcional: terminar la ejecución si faltan variables críticas
    # sys.exit(1)
from flask import (
    Flask, request, render_template, redirect, url_for,
    send_from_directory, session, flash, send_file
)
import openpyxl
from openpyxl import load_workbook, Workbook
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from pymongo import MongoClient
from flask_mail import Mail, Message
from bson import ObjectId
import logging
logging.basicConfig(filename="/var/www/vhosts/edefrutos2025.xyz/httpdocs/flask_app.log", level=logging.DEBUG, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")

# Forzar el uso del bundle de certificados de certifi
os.environ['SSL_CERT_FILE'] = certifi.where()

# Importación de AWS S3
import boto3
from botocore.exceptions import ClientError

# Configuración de AWS S3
AWS_ACCESS_KEY_ID = os.getenv('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = os.getenv('AWS_SECRET_ACCESS_KEY')
AWS_REGION = os.getenv('AWS_REGION')
S3_BUCKET_NAME = os.getenv('S3_BUCKET_NAME')

# Inicialización del cliente S3
s3_client = boto3.client(
    's3',
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
    region_name=AWS_REGION
)

# -------------------------------------------
# CONFIGURACIÓN FLASK
# -------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", os.urandom(24).hex())

# Carpeta para imágenes del catálogo
app.config["UPLOAD_FOLDER"] = os.path.join(app.root_path, "imagenes_subidas")
if not os.path.exists(app.config["UPLOAD_FOLDER"]):
    os.makedirs(app.config["UPLOAD_FOLDER"])

# Carpeta para hojas de cálculo (tablas)
SPREADSHEET_FOLDER = os.path.join(app.root_path, "spreadsheets")
if not os.path.exists(SPREADSHEET_FOLDER):
    os.makedirs(SPREADSHEET_FOLDER)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def eliminar_archivo_imagen(ruta_imagen):
    """Función auxiliar para eliminar físicamente un archivo de imagen.
    
    Args:
        ruta_imagen: Ruta relativa de la imagen (ej: /imagenes_subidas/imagen.jpg) o URI de S3 (ej: s3://bucket/key)
    
    Returns:
        bool: True si se eliminó correctamente, False en caso contrario
    """
    if not ruta_imagen:
        return False
    
    # Verificar si es una ruta S3
    if ruta_imagen.startswith('s3://'):
        # Extraer el nombre del objeto S3 de la ruta
        try:
            # Formato s3://bucket-name/object-key
            parts = ruta_imagen[5:].split('/', 1)
            if len(parts) != 2:
                print(f"Formato de ruta S3 inválido: {ruta_imagen}")
                return False
            
            bucket_name, object_key = parts
            if bucket_name != S3_BUCKET_NAME:
                print(f"El bucket en la ruta ({bucket_name}) no coincide con el configurado ({S3_BUCKET_NAME})")
                return False
                
            return delete_file_from_s3(object_key)
        except Exception as e:
            print(f"Error al eliminar el archivo de S3 {ruta_imagen}: {str(e)}")
            return False
    else:
        # Comportamiento anterior para archivos locales
        # Convertir ruta relativa a absoluta
        if ruta_imagen.startswith('/'):
            # Eliminar la barra inicial para construir la ruta correctamente
            ruta_imagen = ruta_imagen[1:]
        
        ruta_absoluta = os.path.join(app.root_path, ruta_imagen)
        
        try:
            if os.path.exists(ruta_absoluta):
                os.remove(ruta_absoluta)
                print(f"Archivo eliminado: {ruta_absoluta}")
                return True
            else:
                print(f"El archivo no existe: {ruta_absoluta}")
                return False
        except Exception as e:
            print(f"Error al eliminar el archivo {ruta_absoluta}: {str(e)}")
            return False

def upload_file_to_s3(file_path, object_name=None, max_retries=3, delete_local=True):
    """Sube un archivo a un bucket de S3, verifica que la subida fue exitosa, 
    implementa reintentos en caso de fallo y elimina el archivo local después de la subida
    
    Args:
        file_path: Ruta del archivo a subir
        object_name: Nombre de objeto S3, si es None, se usa el nombre del archivo
        max_retries: Número máximo de intentos de subida (por defecto: 3)
        delete_local: Si True, elimina el archivo local después de una subida exitosa
    
    Returns:
        bool: True si la subida es exitosa, False en caso contrario
    """
    if object_name is None:
        object_name = os.path.basename(file_path)
    
    if not os.path.exists(file_path):
        app.logger.error(f"El archivo {file_path} no existe")
        return False
    
    successful = False
    attempt = 0
    
    while attempt < max_retries and not successful:
        attempt += 1
        app.logger.info(f"Intento {attempt}/{max_retries} - Subiendo archivo a S3: {object_name}")
        
        try:
            # Subir el archivo a S3
            s3_client.upload_file(file_path, S3_BUCKET_NAME, object_name)
            
            # Verificar que el archivo existe en S3
            try:
                s3_client.head_object(Bucket=S3_BUCKET_NAME, Key=object_name)
                successful = True
                app.logger.info(f"✅ Archivo subido y verificado en S3: {object_name}")
            except ClientError as e:
                app.logger.warning(f"⚠️ Archivo subido pero no se pudo verificar en S3: {object_name}. Error: {e}")
                # Si el archivo no se puede verificar, se considera como un error
                successful = False
                
        except ClientError as e:
            app.logger.error(f"❌ Error al subir archivo a S3 (intento {attempt}/{max_retries}): {e}")
            if attempt < max_retries:
                import time
                # Esperar un tiempo incremental entre reintentos (backoff exponencial)
                wait_time = 2 ** attempt
                app.logger.info(f"Esperando {wait_time} segundos antes del siguiente intento...")
                time.sleep(wait_time)
    
    # Si la subida fue exitosa y se solicita eliminar el archivo local
    if successful and delete_local:
        try:
            os.remove(file_path)
            app.logger.info(f"🗑️ Archivo local eliminado después de subida exitosa: {file_path}")
        except Exception as e:
            app.logger.warning(f"⚠️ No se pudo eliminar el archivo local {file_path}: {e}")
    
    return successful

def delete_file_from_s3(object_name):
    """Elimina un archivo de un bucket de S3
    
    Args:
        object_name: Nombre del objeto a eliminar
    
    Returns:
        bool: True si la eliminación es exitosa, False en caso contrario
    """
    try:
        s3_client.delete_object(Bucket=S3_BUCKET_NAME, Key=object_name)
        print(f"Archivo eliminado de S3: {object_name}")
        return True
    except ClientError as e:
        print(f"Error eliminando archivo de S3: {e}")
        return False

def get_s3_url(object_name, expiration=3600):
    """Genera una URL prefirmada para acceder a un objeto de S3
    
    Args:
        object_name: Nombre del objeto S3
        expiration: Tiempo de expiración en segundos (1 hora por defecto)
    
    Returns:
        str: URL generada o None si hay error
    """
    try:
        url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': S3_BUCKET_NAME, 'Key': object_name},
            ExpiresIn=expiration
        )
        return url
    except ClientError as e:
        print(f"Error generando URL prefirmada: {e}")
        return None
# -------------------------------------------
# CONFIGURACIÓN DE EMAIL (Flask-Mail)
# -------------------------------------------
app.config["MAIL_SERVER"] = os.environ.get("MAIL_SERVER")
app.config["MAIL_PORT"] = int(os.environ.get("MAIL_PORT", 587))
app.config["MAIL_USE_TLS"] = os.environ.get("MAIL_USE_TLS", "True") == "True"
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD")
app.config["MAIL_DEFAULT_SENDER"] = (
    os.environ.get("MAIL_DEFAULT_SENDER_NAME"),
    os.environ.get("MAIL_DEFAULT_SENDER_EMAIL")
)
app.config["MAIL_DEBUG"] = os.environ.get("MAIL_DEBUG", "False") == "True"
mail = Mail(app)

# Crear la conexión a MongoDB Atlas
MONGO_URI = os.environ.get("MONGO_URI")
from pymongo.server_api import ServerApi
client = MongoClient(
    MONGO_URI,
    tls=True,
    tlsCAFile=certifi.where(),
    server_api=ServerApi('1')
)

try:
    client.admin.command('ping')
    print("✅ Pingó su implementación. ¡Te conectaste con éxito a MongoDB!")
except Exception as e:
    print("❌ Error al conectar con MongoDB:", e)

# Conectar a la base de datos
db = client["app_catalogojoyero"]

# Mostrar colecciones disponibles
print("\U0001F4CC Colecciones disponibles en MongoDB:", db.list_collection_names())

# Definir la colección específica del catálogo
catalog_collection = db["67b8c24a7fdc72dd4d8703cf"]  # Asegúrate de que el nombre es el correcto
users_collection = db["users"]
resets_collection = db["password_resets"]
spreadsheets_collection = db["spreadsheets"]

# Verificar si la colección tiene datos
registros = list(catalog_collection.find())
# Comentado para evitar imprimir todos los documentos
# print("\U0001F4CC Documentos en la colección:")
# for doc in registros:
#     print(doc)
print(f"\U0001F4CC Total de registros en la colección: {len(registros)}")


# Insertar un documento de prueba si la colección está vacía
if len(registros) == 0:
    doc_prueba = {"test": "Conexión funcionando"}
    catalog_collection.insert_one(doc_prueba)
    print("✅ Se insertó un documento de prueba.")

# Comentado para evitar la impresión extensa de documentos
# print("\U0001F4CC Registros después de la prueba:", list(catalog_collection.find()))

# -------------------------------------------
# FUNCIONES AUXILIARES PARA HOJAS DE CÁLCULO
# -------------------------------------------
def leer_datos_excel(filename):
    if not os.path.exists(filename):
        return []

    wb = load_workbook(filename, read_only=True)
    hoja = wb.active
    data = []

    headers = [cell.value for cell in hoja[1]]

    for row in hoja.iter_rows(min_row=2, values_only=True):
        registro = {headers[i]: row[i] for i in range(len(headers))}

        # Asegurar que 'Imagenes' sea una lista
        if "Imagenes" in registro and isinstance(registro["Imagenes"], str):
            registro["Imagenes"] = registro["Imagenes"].split(", ")
        elif "Imagenes" in registro:
            registro["Imagenes"] = []

        data.append(registro)

    wb.close()
    return data

def escribir_datos_excel(data, filename):
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Datos"

    headers = session.get("selected_headers", ["Número", "Descripción", "Peso", "Valor", "Imagenes"])

    if "Número" not in headers:
        headers.insert(0, "Número")

    hoja.append(headers)

    for item in data:
        fila = []
        for header in headers:
            valor = item.get(header, "")

            if header == "Imagenes" and isinstance(valor, list):
                valor = ", ".join(valor)

            fila.append(valor)

        hoja.append(fila)

    wb.save(filename)
    wb.close()

def get_current_spreadsheet():
    filename = session.get("selected_table")
    if not filename:
        return None
    return os.path.join(SPREADSHEET_FOLDER, filename)

# -------------------------------------------
# NUEVA RUTA: PÁGINA DE BIENVENIDA (MANUAL DE USO)
# -------------------------------------------
@app.route("/welcome")
def welcome():
    # Si el usuario no ha iniciado sesión, redirige al login
    if "usuario" not in session:
        return redirect(url_for("login"))
    return render_template("welcome.html")

# -------------------------------------------
# RUTAS DE AUTENTICACIÓN
# -------------------------------------------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        nombre = request.form.get("nombre").strip()
        email = request.form.get("email").strip().lower()
        password = request.form.get("password").strip()
        if users_collection.find_one({"email": email}):
            return "Error: Ese email ya está registrado. <a href='/register'>Volver</a>"
        hashed = generate_password_hash(password)
        nuevo_usuario = {"nombre": nombre, "email": email, "password": hashed}
        users_collection.insert_one(nuevo_usuario)
        return "Registro exitoso. <a href='/login'>Iniciar Sesión</a>"
    else:
        return render_template("register.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        login_input = request.form.get("login_input").strip()
        password = request.form.get("password").strip()
        usuario = users_collection.find_one({
            "$or": [
                {"nombre": {"$regex": f"^{login_input}$", "$options": "i"}},
                {"email": {"$regex": f"^{login_input}$", "$options": "i"}}
            ]
        })
        if not usuario:
            return "Error: Usuario no encontrado. <a href='/login'>Reintentar</a>"
        if check_password_hash(usuario["password"], password):
            session["usuario"] = usuario["nombre"]
            return redirect(url_for("home"))
        else:
            return "Error: Contraseña incorrecta. <a href='/login'>Reintentar</a>"
    else:
        return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/recover")
def recover_redirect():
    return redirect(url_for("forgot_password"))

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        usuario_input = request.form.get("usuario").strip()
        user = users_collection.find_one({
            "$or": [
                {"email": {"$regex": f"^{usuario_input}$", "$options": "i"}},
                {"nombre": {"$regex": f"^{usuario_input}$", "$options": "i"}}
            ]
        })
        if not user:
            return "No se encontró ningún usuario con ese nombre o email. <a href='/forgot-password'>Volver</a>"
        token = secrets.token_urlsafe(32)
        expires_at = datetime.utcnow() + timedelta(minutes=30)
        resets_collection.insert_one({
            "user_id": user["_id"],
            "token": token,
            "expires_at": expires_at,
            "used": False
        })
        reset_link = url_for("reset_password", token=token, _external=True)
        msg = Message("Recuperación de contraseña", recipients=[user["email"]])
        msg.body = (f"Hola {user['nombre']},\n\nPara restablecer tu contraseña, haz clic en el siguiente enlace:\n"
                    f"{reset_link}\n\nEste enlace caduca en 30 minutos.")
        mail.send(msg)
        return "Se ha enviado un enlace de recuperación a tu email. <a href='/login'>Inicia Sesión</a>"
    else:
        return render_template("forgot_password.html")

@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    token = request.args.get("token") or request.form.get("token")
    if not token:
        return "Token no proporcionado."
    reset_info = resets_collection.find_one({"token": token})
    if not reset_info:
        return "Token inválido o inexistente."
    if reset_info["used"]:
        return "Este token ya ha sido utilizado."
    if datetime.utcnow() > reset_info["expires_at"]:
        return "Token caducado."
    if request.method == "POST":
        new_pass = request.form.get("password").strip()
        hashed = generate_password_hash(new_pass)
        user_id = reset_info["user_id"]
        users_collection.update_one({"_id": user_id}, {"$set": {"password": hashed}})
        resets_collection.update_one({"_id": reset_info["_id"]}, {"$set": {"used": True}})
        return "Contraseña actualizada con éxito. <a href='/login'>Inicia Sesión</a>"
    else:
        return render_template("reset_password_form.html", token=token)

# -------------------------------------------
# RUTAS PARA GESTIÓN DE TABLAS (SpreadSheets)
# -------------------------------------------
@app.route("/")
def home():
    if "usuario" not in session:
        return render_template("welcome.html")  # Mostrar página de bienvenida si no hay sesión iniciada

    if "selected_table" in session:
        return redirect(url_for("catalog"))
    else:
        return redirect(url_for("tables"))

@app.route("/tables", methods=["GET", "POST"])
def tables():
    try:
        if "usuario" not in session:
            return redirect(url_for("login"))

        owner = session["usuario"]

        if request.method == "POST":
            try:
                table_name = request.form.get("table_name", "").strip()
                import_file = request.files.get("import_table")

                if import_file and import_file.filename != "":
                    try:
                        # Importar el Excel
                        filename = secure_filename(import_file.filename)
                        filepath = os.path.join(SPREADSHEET_FOLDER, filename)
                        import_file.save(filepath)
                        
                        try:
                            # Leer los encabezados del Excel importado
                            wb = openpyxl.load_workbook(filepath)
                            hoja = wb.active
                            headers = next(hoja.iter_rows(min_row=1, max_row=1, values_only=True))
                            wb.close()
                        except Exception as e:
                            app.logger.error(f"Error al leer el archivo Excel: {str(e)}")
                            flash("Error al leer el archivo Excel. Verifique que el formato sea correcto.", "error")
                            return redirect(url_for("tables"))
                        
                        # Verificar que los encabezados no estén vacíos
                        if not headers or all(header is None for header in headers):
                            flash("El archivo Excel importado no contiene encabezados válidos.", "error")
                            return redirect(url_for("tables"))

                        try:
                            # Guardar info en MongoDB
                            spreadsheets_collection.insert_one({
                                "owner": owner,
                                "name": table_name,
                                "filename": filename,
                                "headers": headers,
                                "created_at": datetime.utcnow()
                            })
                        except Exception as e:
                            app.logger.error(f"Error al guardar en MongoDB: {str(e)}")
                            flash("Error al guardar la información de la tabla en la base de datos.", "error")
                            return redirect(url_for("tables"))
                    except Exception as e:
                        app.logger.error(f"Error al guardar el archivo subido: {str(e)}")
                        flash("Error al procesar el archivo subido. Inténtelo de nuevo.", "error")
                        return redirect(url_for("tables"))
                else:
                    try:
                        # Caso en que no sube archivo (creas uno nuevo con encabezados)
                        headers_str = request.form.get("table_headers", "").strip()
                        if not headers_str:
                            headers = ["Número", "Descripción", "Peso", "Valor"]  # Por defecto
                        else:
                            headers = [h.strip() for h in headers_str.split(",") if h.strip()]

                        # Verificar que los encabezados no estén vacíos
                        if not headers:
                            flash("Debe proporcionar al menos un encabezado válido.", "error")
                            return redirect(url_for("tables"))
                        
                        file_id = secrets.token_hex(8)
                        filename = f"table_{file_id}.xlsx"
                        filepath = os.path.join(SPREADSHEET_FOLDER, filename)
                        
                        try:
                            wb = Workbook()
                            hoja = wb.active
                            hoja.append(headers)
                            wb.save(filepath)
                            wb.close()
                        except Exception as e:
                            app.logger.error(f"Error al crear el archivo Excel: {str(e)}")
                            flash("Error al crear el archivo Excel. Inténtelo de nuevo.", "error")
                            return redirect(url_for("tables"))

                        try:
                            # Guardar info en MongoDB
                            spreadsheets_collection.insert_one({
                                "owner": owner,
                                "name": table_name,
                                "filename": filename,
                                "headers": headers,
                                "created_at": datetime.utcnow()
                            })
                        except Exception as e:
                            app.logger.error(f"Error al guardar en MongoDB: {str(e)}")
                            flash("Error al guardar la información de la tabla en la base de datos.", "error")
                            return redirect(url_for("tables"))
                    except Exception as e:
                        app.logger.error(f"Error al crear nueva tabla: {str(e)}")
                        flash("Error al crear la nueva tabla. Inténtelo de nuevo.", "error")
                        return redirect(url_for("tables"))
                
                try:
                    # Almacenar los encabezados en la sesión para su uso posterior
                    session["selected_headers"] = headers
                except Exception as e:
                    app.logger.warning(f"Error al guardar encabezados en sesión: {str(e)}")
                    # No es crítico, continuamos
                return redirect(url_for("tables"))
            except Exception as e:
                app.logger.error(f"Error general en el procesamiento POST de tablas: {str(e)}")
                flash("Error al procesar la solicitud. Inténtelo de nuevo.", "error")
                return redirect(url_for("tables"))

        try:
            # GET: mostrar las tablas existentes
            todas_las_tablas = list(spreadsheets_collection.find({"owner": session["usuario"]}))
            return render_template("tables.html", tables=todas_las_tablas)
        except Exception as e:
            app.logger.error(f"Error al consultar tablas en MongoDB: {str(e)}")
            flash("Error al cargar las tablas. Inténtelo de nuevo más tarde.", "error")
            return render_template("tables.html", tables=[])
    except Exception as e:
        app.logger.error(f"Error crítico en la función tables(): {str(e)}")
        flash("Ha ocurrido un error inesperado. Por favor, inténtelo de nuevo.", "error")
        return redirect(url_for("welcome"))

@app.route("/select_table/<table_id>")
def select_table(table_id):
    if "usuario" not in session:
        return redirect(url_for("login"))

    table = spreadsheets_collection.find_one({"_id": ObjectId(table_id)})

    if not table:
        flash("Tabla no encontrada.", "error")
        return redirect(url_for("tables"))

    session["selected_table"] = table["filename"]
    session["selected_table_id"] = str(table["_id"])  # Almacenar el ID de la tabla
    session["selected_table_name"] = table["name"]  # Almacenar el nombre de la tabla

    return redirect(url_for("catalog"))


# -------------------------------------------
# RUTAS DEL CATÁLOGO (Excel e imágenes) para la tabla seleccionada
# -------------------------------------------

@app.route("/catalog", methods=["GET", "POST"])
def catalog():
    if "usuario" not in session:
        return redirect(url_for("welcome"))

    if "selected_table" not in session:
        flash("Por favor, seleccione una tabla primero.", "warning")
        return redirect(url_for("tables"))

    selected_table = session["selected_table"]
    table_info = spreadsheets_collection.find_one({"filename": selected_table})

    if not table_info:
        flash("La tabla seleccionada no existe.", "error")
        return redirect(url_for("tables"))

    headers = table_info.get("headers", [])
    if not headers:
        flash("La tabla seleccionada no tiene encabezados definidos.", "error")
        return redirect(url_for("tables"))

    # Guardar los encabezados en la sesión
    session["selected_headers"] = headers

    # Obtener registros de MongoDB para la tabla seleccionada
    # Asegurarse de que el campo "Número" es entero antes de ordenar
    pipeline = [
        {"$match": {"table": selected_table}},
        {"$addFields": {"NumeroOrdenacion": {"$toInt": {"$ifNull": [{"$toInt": "$Número"}, "$Número"]}}}},
        {"$sort": {"NumeroOrdenacion": 1}}
    ]
    registros = list(catalog_collection.aggregate(pipeline))

    if request.method == "POST":
        form_data = {k.strip(): v.strip() for k, v in request.form.items()}

        # Verificar si existe "Número" o el primer encabezado como identificador
        id_field = headers[0]
        if id_field not in form_data or not form_data[id_field]:
            return render_template("index.html", data=registros, headers=headers, error_message=f"Error: Sin {id_field}.")

        # Verificar si el identificador ya existe en esta tabla
        if any(item.get(id_field) == form_data[id_field] for item in registros):
            return render_template("index.html", data=registros, headers=headers, error_message=f"Error: Ese {id_field} ya existe.")

        # Construir el nuevo registro
        nuevo_registro = {
            "Número": len(registros) + 1,
            "table": selected_table
        }
        
        # Agregar los campos del formulario
        for header in headers:
            safe_header = header.replace(" ", "_").replace(".", "_")
            nuevo_registro[safe_header] = form_data.get(header, "").strip()

        # Manejo de imágenes
        # Manejo de imágenes
        files = request.files.getlist("imagenes")
        rutas_imagenes = [None, None, None]
        
        for i, file in enumerate(files[:3]):
            if file and allowed_file(file.filename):
                # Generar un nombre único con timestamp y uuid para evitar colisiones
                original_filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
                unique_id = secrets.token_hex(4)
                extension = os.path.splitext(original_filename)[1]
                unique_filename = f"{timestamp}_{unique_id}{extension}"
                
                app.logger.info(f"Procesando imagen {i+1} para nuevo registro: {unique_filename}")
                
                # Guardar temporalmente el archivo
                temp_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
                file.save(temp_path)
                
                try:
                    # Subir a S3 con nuestra función mejorada que verifica la subida y elimina archivos locales
                    if upload_file_to_s3(temp_path, unique_filename, delete_local=True):
                        # Si la carga a S3 es exitosa, guardamos la referencia a S3
                        s3_url = f"s3://{S3_BUCKET_NAME}/{unique_filename}"
                        rutas_imagenes[i] = s3_url
                        app.logger.info(f"Imagen {i+1} subida exitosamente a S3: {s3_url}")
                    else:
                        app.logger.error(f"Falló la subida a S3 para imagen {i+1}: {unique_filename}")
                        # No almacenamos la imagen localmente, solo registramos el error
                except Exception as e:
                    app.logger.error(f"Error al procesar imagen {i+1}: {str(e)}")
                    # Asegurar que el archivo temporal se elimina incluso si hay error
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                        app.logger.info(f"Archivo temporal eliminado: {temp_path}")
        # Asignar las rutas de imágenes al nuevo registro
        nuevo_registro["Imagenes"] = rutas_imagenes

        # Insertar en MongoDB
        catalog_collection.insert_one(nuevo_registro)

        return redirect(url_for("catalog"))

    # GET: mostrar los registros existentes
    return render_template("index.html", data=registros, headers=headers)

@app.route("/editar/<id>", methods=["GET", "POST"])
def editar(id):
    if "usuario" not in session:
        return redirect(url_for("login"))

    if "selected_table" not in session:
        flash("Por favor, seleccione una tabla primero.", "warning")
        return redirect(url_for("tables"))

    selected_table = session["selected_table"]
    table_info = spreadsheets_collection.find_one({"filename": selected_table})

    if not table_info:
        flash("La tabla seleccionada no existe.", "error")
        return redirect(url_for("tables"))

    headers = table_info.get("headers", [])
    if not headers:
        flash("La tabla seleccionada no tiene encabezados definidos.", "error")
        return redirect(url_for("tables"))

    id_field = headers[0]
    safe_id_field = id_field.replace(" ", "_").replace(".", "_")

    # Obtenemos el registro desde MongoDB - Intentamos primero con "Número" para tablas manuales
    registro = None
    encontrado_por_numero = False  # Variable para rastrear cómo se encontró el registro
    
    try:
        # Intenta buscar por "Número" convertido a entero
        num_id = int(id)
        registro = catalog_collection.find_one({"Número": num_id, "table": selected_table})
        if registro:
            encontrado_por_numero = True  # Marcamos que se encontró por Número
    except (ValueError, TypeError):
        # Si la conversión falla, continúa con la búsqueda normal
        pass
        
    # Si no encontramos el registro por "Número", intentamos con el encabezado original
    if not registro:
        registro = catalog_collection.find_one({safe_id_field: id, "table": selected_table})
        
    if not registro:
        flash(f"No existe el registro con ID {id} en la tabla seleccionada.", "error")
        return redirect(url_for("catalog"))

    if request.method == "GET":
        # Filtrar los headers para excluir 'Imagenes' del formulario principal
        headers_form = [h for h in headers if h != "Imagenes"]
        return render_template("editar.html", 
                             registro=registro, 
                             headers=headers_form,
                             imagenes_actuales=registro.get("Imagenes", [None, None, None]))

    # POST: Guardar cambios
    # POST: Guardar cambios
    if request.form.get("delete_record") == "on":
        # Primero, eliminar físicamente las imágenes asociadas al registro
        rutas_imagenes = registro.get("Imagenes", [])
        for ruta in rutas_imagenes:
            if ruta:  # Solo intentar eliminar si la ruta no es None
                eliminar_archivo_imagen(ruta)
                
        # Ahora eliminar el registro de la base de datos
        # Usar el mismo criterio por el que se encontró el registro originalmente
        if encontrado_por_numero:
            # Si se encontró por "Número", eliminamos usando ese campo
            result = catalog_collection.delete_one({"Número": int(id), "table": selected_table})
        else:
            # Si se encontró por safe_id_field, eliminamos usando ese campo
            result = catalog_collection.delete_one({safe_id_field: id, "table": selected_table})
            
        if result.deleted_count > 0:
            # Renumerar registros para evitar huecos en la numeración
            renumerar_registros(selected_table)
            flash("Registro y sus imágenes eliminados exitosamente.", "success")
        else:
            flash("No se pudo eliminar el registro.", "error")
        return redirect(url_for("catalog"))
    try:
        # Actualizar campos (excluyendo imágenes)
        update_data = {}
        
        # Manejar campos especiales primero
        update_data["Número"] = registro["Número"]  # Mantener el número original
        update_data["table"] = selected_table

        # Actualizar el resto de campos desde el formulario
        for header in headers:
            if header != "Imagenes" and header != "Número":
                form_value = request.form.get(header, "").strip()
                # Usar punto para campos con espacios en MongoDB
                safe_header = header.replace(" ", "_").replace(".", "_")
                update_data[safe_header] = form_value

        # Manejo de imágenes
        # Manejo de imágenes
        rutas_imagenes = registro.get("Imagenes", [None, None, None])
        
        # Procesar cada campo de imagen por separado
        # Imagen 1
        imagen1 = request.files.get("imagen1")
        if imagen1 and imagen1.filename and allowed_file(imagen1.filename):
            # Generar un nombre único con timestamp y uuid para evitar colisiones
            original_filename = secure_filename(imagen1.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
            unique_id = secrets.token_hex(4)
            extension = os.path.splitext(original_filename)[1]
            unique_filename = f"{timestamp}_{unique_id}{extension}"
            
            app.logger.info(f"Procesando imagen 1 para actualización: {unique_filename}")
            
            # Guardar temporalmente el archivo
            temp_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
            imagen1.save(temp_path)
            
            try:
                # Subir a S3 con nuestra función mejorada que verifica la subida y elimina archivos locales
                if upload_file_to_s3(temp_path, unique_filename, delete_local=True):
                    # Si la carga a S3 es exitosa, guardamos la referencia a S3
                    s3_url = f"s3://{S3_BUCKET_NAME}/{unique_filename}"
                    rutas_imagenes[0] = s3_url
                    app.logger.info(f"Imagen 1 subida exitosamente a S3: {s3_url}")
                else:
                    app.logger.error(f"Falló la subida a S3 para imagen 1: {unique_filename}")
                    # No almacenamos la imagen localmente, solo registramos el error
            except Exception as e:
                app.logger.error(f"Error al procesar imagen 1: {str(e)}")
                # Asegurar que el archivo temporal se elimina incluso si hay error
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                    app.logger.info(f"Archivo temporal eliminado: {temp_path}")
            
        # Imagen 2
        imagen2 = request.files.get("imagen2")
        if imagen2 and imagen2.filename and allowed_file(imagen2.filename):
            # Generar un nombre único con timestamp y uuid para evitar colisiones
            original_filename = secure_filename(imagen2.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
            unique_id = secrets.token_hex(4)
            extension = os.path.splitext(original_filename)[1]
            unique_filename = f"{timestamp}_{unique_id}{extension}"
            
            app.logger.info(f"Procesando imagen 2 para actualización: {unique_filename}")
            
            # Guardar temporalmente el archivo
            temp_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
            imagen2.save(temp_path)
            
            try:
                # Subir a S3 con nuestra función mejorada que verifica la subida y elimina archivos locales
                if upload_file_to_s3(temp_path, unique_filename, delete_local=True):
                    # Si la carga a S3 es exitosa, guardamos la referencia a S3
                    s3_url = f"s3://{S3_BUCKET_NAME}/{unique_filename}"
                    rutas_imagenes[1] = s3_url
                    app.logger.info(f"Imagen 2 subida exitosamente a S3: {s3_url}")
                else:
                    app.logger.error(f"Falló la subida a S3 para imagen 2: {unique_filename}")
                    # No almacenamos la imagen localmente, solo registramos el error
            except Exception as e:
                app.logger.error(f"Error al procesar imagen 2: {str(e)}")
                # Asegurar que el archivo temporal se elimina incluso si hay error
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                    app.logger.info(f"Archivo temporal eliminado: {temp_path}")
            
        # Imagen 3
        # Imagen 3
        imagen3 = request.files.get("imagen3")
        if imagen3 and imagen3.filename and allowed_file(imagen3.filename):
            # Generar un nombre único con timestamp y uuid para evitar colisiones
            original_filename = secure_filename(imagen3.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
            unique_id = secrets.token_hex(4)
            extension = os.path.splitext(original_filename)[1]
            unique_filename = f"{timestamp}_{unique_id}{extension}"
            
            app.logger.info(f"Procesando imagen 3 para actualización: {unique_filename}")
            
            # Guardar temporalmente el archivo
            temp_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
            imagen3.save(temp_path)
            
            try:
                # Subir a S3 con nuestra función mejorada que verifica la subida y elimina archivos locales
                if upload_file_to_s3(temp_path, unique_filename, delete_local=True):
                    # Si la carga a S3 es exitosa, guardamos la referencia a S3
                    s3_url = f"s3://{S3_BUCKET_NAME}/{unique_filename}"
                    rutas_imagenes[2] = s3_url
                    app.logger.info(f"Imagen 3 subida exitosamente a S3: {s3_url}")
                else:
                    app.logger.error(f"Falló la subida a S3 para imagen 3: {unique_filename}")
                    # No almacenamos la imagen localmente, solo registramos el error
            except Exception as e:
                app.logger.error(f"Error al procesar imagen 3: {str(e)}")
                # Asegurar que el archivo temporal se elimina incluso si hay error
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                    app.logger.info(f"Archivo temporal eliminado: {temp_path}")
        # Manejar eliminación de imágenes
        for i in range(3):
            if request.form.get(f"remove_img{i+1}") == "on":
                # Eliminar físicamente el archivo antes de actualizar la ruta en la base de datos
                ruta_actual = rutas_imagenes[i]
                if ruta_actual:  # Solo intentar eliminar si hay una imagen
                    eliminar_archivo_imagen(ruta_actual)
                rutas_imagenes[i] = None
        update_data["Imagenes"] = rutas_imagenes
        # Actualizar en MongoDB usando replace_one en lugar de update_one
        # Usar el mismo criterio por el que se encontró el registro originalmente
        if encontrado_por_numero:
            # Si se encontró por "Número", actualizamos usando ese campo
            result = catalog_collection.replace_one(
                {"Número": int(id), "table": selected_table},
                update_data
            )
        else:
            # Si se encontró por safe_id_field, actualizamos usando ese campo
            result = catalog_collection.replace_one(
                {safe_id_field: id, "table": selected_table},
                update_data
            )

        if result.modified_count > 0:
            flash("Registro actualizado exitosamente.", "success")
        else:
            flash("No se detectaron cambios en el registro.", "info")

    except Exception as e:
        flash(f"Error al actualizar el registro: {str(e)}", "error")
        print(f"Error en la actualización: {str(e)}")  # Para debugging

    return redirect(url_for("catalog"))

@app.route("/delete_table/<table_id>", methods=["POST"])
def delete_table(table_id):
    if "usuario" not in session:
        return redirect(url_for("login"))
    # Buscamos la tabla en la colección, asegurándonos que el usuario actual es el propietario.
    table = spreadsheets_collection.find_one({"_id": ObjectId(table_id), "owner": session["usuario"]})
    if not table:
        flash("Tabla no encontrada o no tienes permiso para eliminarla.", "error")
        return redirect(url_for("tables"))
    
    # Construir la ruta absoluta del archivo Excel
    file_path = os.path.join(SPREADSHEET_FOLDER, table["filename"])
    # Si el archivo existe, lo eliminamos del sistema de archivos.
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            flash(f"Error al eliminar el archivo: {e}", "error")
            return redirect(url_for("tables"))
    
    # Eliminamos el documento de la colección en MongoDB
    spreadsheets_collection.delete_one({"_id": ObjectId(table_id)})
    
    # Si la tabla eliminada era la seleccionada en sesión, la removemos de la sesión.
    if session.get("selected_table") == table["filename"]:
        session.pop("selected_table", None)
    
    flash("Tabla eliminada exitosamente.", "success")
    return redirect(url_for("tables"))

@app.route("/descargar-excel")
def descargar_excel():
    if "usuario" not in session:
        return redirect(url_for("login"))
    spreadsheet_path = get_current_spreadsheet()
    if not spreadsheet_path or not os.path.exists(spreadsheet_path):
        return "El Excel no existe aún."
    temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    with zipfile.ZipFile(temp_zip.name, "w") as zf:
        zf.write(spreadsheet_path, arcname=os.path.basename(spreadsheet_path))
        data = leer_datos_excel(spreadsheet_path)
        image_paths = set()
        for row in data:
            for ruta in row["imagenes"]:
                if ruta:
                    if ruta.startswith('s3://'):
                        # Para imágenes en S3, intentar descargarlas al archivo temporal para incluirlas en el ZIP
                        try:
                            s3_parts = ruta[5:].split('/', 1)
                            if len(s3_parts) == 2:
                                bucket_name, object_key = s3_parts
                                if bucket_name == S3_BUCKET_NAME:
                                    filename = os.path.basename(object_key)
                                    temp_download_path = os.path.join(tempfile.gettempdir(), filename)
                                    s3_client.download_file(S3_BUCKET_NAME, object_key, temp_download_path)
                                    # Agregar la ruta temporal a la lista de imágenes para incluir en el ZIP
                                    image_paths.add(temp_download_path)
                        except Exception as e:
                            print(f"Error al descargar imagen de S3: {str(e)}")
                    else:
                        # Comportamiento anterior para archivos locales
                        absolute_path = os.path.join(app.root_path, ruta)
                        if os.path.exists(absolute_path):
                            image_paths.add(absolute_path)
            arcname = os.path.join("imagenes", os.path.basename(img_path))
            zf.write(img_path, arcname=arcname)
    return send_from_directory(directory=os.path.dirname(temp_zip.name),
                               path=os.path.basename(temp_zip.name),
                               as_attachment=True,
                               download_name="catalogo.zip")

from flask import send_from_directory

@app.route("/imagenes_subidas/<filename>")
def uploaded_images(filename):
    # Verificar si se trata de una solicitud a un archivo almacenado en S3
    s3_param = request.args.get('s3')
    if s3_param == 'true':
        # Generar una URL prefirmada para acceder al objeto en S3
        url = get_s3_url(filename)
        if url:
            return redirect(url)
        return "Error al acceder al archivo", 404
    
    # Si no es un archivo S3, servir desde el sistema de archivos local
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

from bson import ObjectId

def convertir_registros(registros):
    for r in registros:
        if "_id" in r and isinstance(r["_id"], ObjectId):
            r["_id"] = str(r["_id"])
    return registros

@app.route("/debug_mongo")
def debug_mongo():
    print("\U0001F4CC Verificando conexión con MongoDB en Flask")

    colecciones = db.list_collection_names()
    print("\U0001F4CC Colecciones disponibles en MongoDB:", colecciones)

    # Obtener documentos pero solo mostrar el conteo
    documentos = list(catalog_collection.find())
    print(f"\U0001F4CC Total de documentos en la colección: {len(documentos)}")
    
    # Convertir ObjectId a string para poder serializar a JSON
    for doc in documentos:
        if "_id" in doc and isinstance(doc["_id"], ObjectId):
            doc["_id"] = str(doc["_id"])

    # Solo devolver las colecciones y el conteo para evitar respuestas extensas
    return {"colecciones": colecciones, "total_documentos": len(documentos)}

@app.route("/insert_test")
def insert_test():
    nuevo_registro = {
        "Número": 2,
        "Descripción": "Collar de prueba",
        "Peso": 10,
        "Valor": 2500,
        "Imagenes": ["/imagenes_subidas/collar_prueba.jpg"]
    }
    
    catalog_collection.insert_one(nuevo_registro)
    return "✅ Registro de prueba insertado en MongoDB"
# -------------------------------------------
# MAIN
# -------------------------------------------
# ... código existente ... (mantenemos todo hasta la función renumerar_registros)
def renumerar_registros(table_name):
    """Renumera todos los registros de una tabla específica en orden secuencial"""
    # Obtener todos los registros ordenados por el campo Número
    registros = list(catalog_collection.find({"table": table_name}).sort("Número", 1))
    
    # Renumerar secuencialmente
    for i, registro in enumerate(registros, 1):
        # Solo actualizar si el número ha cambiado
        if registro.get("Número") != i:
            catalog_collection.update_one(
                {"_id": registro["_id"]},
                {"$set": {"Número": i}}
            )
    
    return len(registros)

@app.route("/renumerar/<table_name>")
def renumerar(table_name):
    """Renumera los registros de una tabla específica y redirecciona al catálogo"""
    if "usuario" not in session:
        return redirect(url_for("welcome"))
    
    try:
        total = renumerar_registros(table_name)
        flash(f"Se han renumerado {total} registros correctamente.", "success")
    except Exception as e:
        flash(f"Error al renumerar registros: {str(e)}", "error")
    
    # Establecer la tabla seleccionada para la redirección
    session["selected_table"] = table_name
    return redirect(url_for("catalog"))


if __name__ == "__main__":
    app.run(debug=True)
